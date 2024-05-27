# 此文件用于将其他数据集的数据转化成【细胞：基因表达】形式，并过滤掉异常细胞和基因
import pickle
import random
import sys

import math
import pandas as pd
import openpyxl
import os
import zipfile
import gzip
import  datasets
from datasets import Dataset

def gzip_file(input_path, output_path):
    # 将对应文件原地解压缩
    with gzip.open(input_path, 'rb') as f_in:
        if(os.path.exists(output_path)):
                print(output_path+' 已经存在')
        else:
            with open(output_path, 'wb') as f_out:
                try:
                    f_out.write(f_in.read())
                except:
                    print('err '+input_path)
    print(input_path+' unzip success')
def traverse_files(root_path):
    # 遍历路径下所有的.gz文件
    file_paths=[]
    for root, dirs, files in os.walk(root_path):
        for file in files:
            if(file.endswith('.gz')):
                file_path = os.path.join(root, file)
                file_paths.append(file_path)
    return  file_paths
def gzip_all(root_path):
    gzfiles=traverse_files(root_path)
    print(len(gzfiles))
    for gzfile in gzfiles:
        new_path=gzfile.replace('.gz', '')
        new_path=new_path.split('/')
        new_path=new_path[-1]
        new_file_path='./dataSets/dataSets_filtered_unzipped/'+new_path
        gzip_file(gzfile,new_file_path)

    pass

def getID2geneDict(GPL_id):
    datapath='./dataSets/GPL_map_files/'+GPL_id+'_map.xlsx'
    # 此函数用于读取探头ID转化成gene的文件，返回一个字典
    geneDict = {}
    # 读取xlsx文件
    data = pd.read_excel(datapath)
    data_value=data.values
    data_indexs=data.keys()
    # 由于不同GPL平台的数据文件结构都不同，对不同GPL使用不同的处理过程
    if(GPL_id=='GPL6844'):
        id_index=2
        gene_index=44
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]]=item[gene_index]
        print(geneDict)
    if(GPL_id=='GPL6871'):
        id_index = 0
        gene_index = 2
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL10652'):
        id_index = 1
        gene_index = 0
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL14150'):
        id_index = 0
        gene_index = 3
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL14991'):
        id_index = 0
        gene_index = 3
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL15192'): #sequence
        id_index = 0
        gene_index = 1
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL15193'): #sequence
        id_index = 0
        gene_index = 1
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL15553'): # 此文件为空
        id_index = 0
        gene_index = 0
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL18464'):
        id_index = 0
        gene_index = 1
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL18465'):
        id_index = 0
        gene_index = 1
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL20150'):# 此文件为空
        id_index = 0
        gene_index = 0
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL21495'): # 未知spot_id
        id_index = 0
        gene_index = 2
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL21512'): # sequence
        id_index = 0
        gene_index = 1
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL21689'):# 此文件为空
        id_index = 0
        gene_index = 2
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL23660'):# sequence
        gene_index = 3
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL28515'):# 此文件为空
        id_index = 0
        gene_index = 3
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    if (GPL_id == 'GPL33372'):# 此文件为空
        id_index = 0
        gene_index = 3
        print(data_indexs)
        for item in data_value:
            geneDict[item[id_index]] = item[gene_index]
        print(geneDict)
    return geneDict
    pass
# getID2geneDict('GPL21512')
def datatreat(gse_id,filename='none'):
    # 对指定文件进行处理,返回经过过滤的细胞基因数据
    data_cells_all=[]
    cell_ids=[]
    if(gse_id=='GSE11503'):
        with open('./dataSets/dataSets_filtered_unzipped/'+filename,'r') as f:
            data_cells = f.readlines()
            print(len(data_cells))
            gene_id=0
            data_cell = data_cells[0]
            data_cell=data_cell.strip('\n')
            data_cell=data_cell.split('\t')
            gene_id = data_cell.index('gene description')
            print(data_cell)
            # 'Field', 'Meta Row', 'Meta Column', 'Row', 'Column', 'accession number', 'clone', 'swiss prot id', 'gene description', 'blast score'
            gene_id=data_cell.index('Row')
            for i in range(1,len(data_cells)):
                data_cell = data_cells[i]
                data_cell = data_cell.strip('\n')
                data_cell = data_cell.split('\t')
                # cell_id=data_cell[9]   # 'blast score'
                # cell_id2 = data_cell[5] # accession number
                cell_id3 = data_cell[1]  # Meta Row
                gene_description=data_cell[5]
                if(cell_id3 in cell_ids):
                    data_cells_all[cell_ids.index(cell_id3)].append(gene_description)
                else:
                    cell_ids.append(cell_id3)
                    data_cells_all.append([gene_description])
            print(data_cells_all)
            print(cell_ids)
        return data_cells_all
    if (gse_id == 'GSE100818'):
        # 实在看不出来格式，跳过
        return data_cells_all
        # with open('./dataSets/dataSets_filtered_unzipped/'+filename,'r') as f:
        #     data_cells = f.readlines()
        #     print(data_cells[0])
        #     index=data_cells[5].split('\t')
        #     data=data_cells[6].split('\t')
        #
        #     result = {k: v for k, v in zip(index, data)}
        #     print(result)
        # pass
    if (gse_id == 'GSE103633'):
        # 基因测序实验，和本实验没关系，跳过
        return data_cells_all
        # with open('./dataSets/dataSets_filtered_unzipped/'+filename,'r') as f:
        #     data_cells = f.readlines()
        #     # for i in range(len(data_cells)):
        #     for i in range(0,3):
        #         data_cell = data_cells[i]
        #         data_cell = data_cell.strip('\n')
        #         data_cell = data_cell.split('\t')
        #         print(data_cell)
    if (gse_id == 'GSE146685'):
        # 基因测序实验，和本实验没关系，跳过
        return data_cells_all
        # with open('./dataSets/dataSets_filtered_unzipped/'+filename,'r') as f:
        #     data_cells = f.readlines()
        #     print(data_cells)
    if (gse_id == 'GSE150259'):
        # 基因测序实验，和本实验没关系，跳过
        # with open('./dataSets/dataSets_filtered_unzipped/'+filename,'r') as f:
        #     data_cells = f.readlines()
        #         # for i in range(len(data_cells)):
        #     for i in range(0,3):
        #         data_cell = data_cells[i]
        #         data_cell = data_cell.strip('\n')
        #         data_cell = data_cell.split('\t')
        #         print(data_cell)
        return data_cells_all
    if (gse_id == 'GSE157860'):
        # 打开xlsx文件
        workbook = openpyxl.load_workbook('./dataSets/dataSets_filtered_unzipped/'+filename)
        # 获取第一个工作表
        sheet = workbook.active
        cols = sheet.max_column
        data_cells_all ={}
        # 遍历每一行数据
        for row in sheet.iter_rows(values_only=True):
            # 输出每一行的数据
            row_data=[cell for cell in row]
            gene_id=row_data[0]
            if(gene_id =='name'):
                for i in range(1,cols):
                    data_cells_all[row_data[i]]=[]
                    cell_ids.append(row_data[i])
            else:
                for i in range(1,cols):
                    if(row_data[i]!=0):
                        data_cells_all[cell_ids[i-1]].append(gene_id)
        # 关闭xlsx文件
        workbook.close()
        return data_cells_all
    if (gse_id == 'GSE158706'):
        # 基因测序与本实验无关
        return data_cells_all
    if (gse_id == 'GSE179290'):
        # GSM5413716	咽后伤口碎片，0h
        # GSM5413717	咽后伤口碎片，6h_calcein
        # GSM5413718	咽后伤口碎片，18h_anterior_faceing
        # GSM5413719	咽后伤口碎片，18h_posterior_faceing
        # GSM5413720	咽后伤口碎片，18h_anterior_faceing_unc22
        # barcodes文件为基因序列，features为基因id,matrix为细胞基因表达数据
        gse_files=['GSM5413717_1286_barcodes.tsv', 'GSM5413718_678_barcodes.tsv', 'GSM5413716_677_features.tsv',
         'GSM5413717_1286_matrix.mtx', 'GSM5413719_679_barcodes.tsv', 'GSM5413718_678_matrix.mtx',
         'GSM5413720_681_features.tsv', 'GSM5413716_677_barcodes.tsv', 'GSM5413716_677_matrix.mtx',
         'GSM5413719_679_matrix.mtx', 'GSM5413717_1286_features.tsv', 'GSM5413720_681_barcodes.tsv',
         'GSM5413720_681_matrix.mtx', 'GSM5413719_679_features.tsv', 'GSM5413718_678_features.tsv']
        data_cells_all={}
        data_cells_all['0h']=[]
        data_cells_all['6h_calcein']=[]
        data_cells_all['18h_anterior_faceing']=[]
        data_cells_all['18h_posterior_faceing']=[]
        data_cells_all['18h_anterior_faceing_unc22']=[]
        cells_num_all=0
        dataSet_list={}
        cell_ids={}
        cell_ids['0h'] = 'GSM5413716_677'
        cell_ids['6h_calcein'] = 'GSM5413717_1286'
        cell_ids['18h_anterior_faceing'] = 'GSM5413718_678'
        cell_ids['18h_posterior_faceing'] = 'GSM5413719_679'
        cell_ids['18h_anterior_faceing_unc22'] = 'GSM5413720_681'
        # 依次对这5轮实验的数据进行处理
        for file_type in data_cells_all.keys():
            data_cell=[]
            file_id= cell_ids[file_type]
            # 获取基因id
            gene_id={}
            with open('./dataSets/dataSets_filtered_unzipped/'+file_id+'_features.tsv','r') as f:
                data_lists=f.readlines()
                i=1
                for data_list in data_lists:
                    data_list=data_list.strip('\n')
                    data_list=data_list.split('\t')
                    gene_id[str(i)]=data_list[0]
                    i+=1
            # 将细胞基因表达矩阵转化成列表
            with open('./dataSets/dataSets_filtered_unzipped/' + file_id + '_matrix.mtx', 'r') as f:
                data_lists = f.readlines()
                data_list=data_lists[2]
                data_list=data_list.strip('\n')
                data_list=data_list.split(' ')
                genes_num=data_list[0]
                cell_num=int(data_list[1])
                data_cell={}
                j=0
                # for i in range(3,len(data_lists)):
                for data_list in data_lists:
                    if(j<=2):
                        j+=1
                        continue
                    j+=1
                    data_list = data_list.strip('\n')
                    data_list = data_list.split(' ')
                    cell_gene=gene_id[data_list[0]]
                    cell_id=data_list[1]
                    if(cell_id in data_cell.keys()):
                        data_cell[cell_id]['cell_genes'].append(cell_gene)
                    else:
                        data_cell[cell_id]={'cell_id':cell_id,'cell_genes':[cell_gene],'file_type':file_type}

            for id in data_cell.keys():
                data_cells_all[file_type].append(data_cell[id])
            cells_num_all += len(data_cells_all[file_type])
            print(cells_num_all)
            dataSet_item=Dataset.from_list(data_cells_all[file_type])
            dataSet_list[file_type]=dataSet_item
        print(gse_id+' end with num:'+str(cells_num_all))
        return dataSet_list
    if (gse_id == 'GSE231548'):
        gse_files=['GSM7290662_rep1_dash_features.tsv', 'GSM7290662_rep1_dash_matrix.mtx', 'GSM7290663_rep2_dash_matrix.mtx', 'GSM7290660_rep2_matrix.mtx',
                   'GSM7290663_rep2_dash_features.tsv', 'GSM7290660_rep2_features.tsv', 'GSM7290665_rep1_noPCR_matrix.mtx', 'GSM7290660_rep2_barcodes.tsv',
                   'GSM7290665_rep1_noPCR_barcodes.tsv', 'GSM7290661_rep3_features.tsv', 'GSM7290666_rep1_5cycles_features.tsv',
                   'GSM7290661_rep3_barcodes.tsv', 'GSM7290667_rep1_10cycles_matrix.mtx', 'GSM7290667_rep1_10cycles_barcodes.tsv',
                   'GSM7290664_rep3_dash_matrix.mtx', 'GSM7290659_rep1_features.tsv', 'GSM7290659_rep1_barcodes.tsv', 'GSM7290664_rep3_dash_barcodes.tsv',
                   'GSM7290664_rep3_dash_features.tsv', 'GSM7290663_rep2_dash_barcodes.tsv', 'GSM7290662_rep1_dash_barcodes.tsv',
                   'GSM7290665_rep1_noPCR_features.tsv', 'GSM7290667_rep1_10cycles_features.tsv', 'GSM7290666_rep1_5cycles_barcodes.tsv',
                   'GSM7290661_rep3_matrix.mtx', 'GSM7290666_rep1_5cycles_matrix.mtx', 'GSM7290659_rep1_matrix.mtx']
        # GSM7290659	未经处理，重复 1，scRNAseq
        # GSM7290660	未经处理，重复 2，scRNAseq
        # GSM7290661	未经处理，重复 3，scRNAseq
        # GSM7290662	DASHed，重复 1，scRNAseq
        # GSM7290663	DASHed，重复 2，scRNAseq
        # GSM7290664	DASHed，重复 3，scRNAseq
        # GSM7290665	noPCR、重复 1、scRNAseq
        # GSM7290666	5 个循环，重复 1 次，scRNAseq
        # GSM7290667	10 个循环，重复 1 次，scRNAseq
        data_cells_all={}
        data_cells_all['GSM7290659']=[]
        data_cells_all['GSM7290660']=[]
        data_cells_all['GSM7290661']=[]
        data_cells_all['GSM7290662'] = []
        data_cells_all['GSM7290663'] = []
        data_cells_all['GSM7290664'] = []
        data_cells_all['GSM7290665'] = []
        data_cells_all['GSM7290666'] = []
        data_cells_all['GSM7290667'] = []
        cells_num_all=0
        cell_ids={}
        cell_ids['GSM7290659']='GSM7290659_rep1'
        cell_ids['GSM7290660']='GSM7290660_rep2'
        cell_ids['GSM7290661']='GSM7290661_rep3'
        cell_ids['GSM7290662']='GSM7290662_rep1_dash'
        cell_ids['GSM7290663']='GSM7290663_rep2_dash'
        cell_ids['GSM7290664']='GSM7290664_rep3_dash'
        cell_ids['GSM7290665']='GSM7290665_rep1_noPCR'
        cell_ids['GSM7290666']='GSM7290666_rep1_5cycles'
        cell_ids['GSM7290667']='GSM7290667_rep1_10cycles'
        for file_type in data_cells_all.keys():
            data_cell = []
            file_id = cell_ids[file_type]
            # 获取基因id
            gene_id = {}
            with open('./dataSets/dataSets_filtered_unzipped/'+file_id+'_features.tsv','r') as f:
                data_lists = f.readlines()
                i = 1
                for data_list in data_lists:
                    data_list = data_list.strip('\n')
                    data_list = data_list.split('\t')
                    gene=data_list[0]
                    gene=gene.split('.')
                    gene_id[str(i)] = gene[0]
                    i += 1
            # 将基因表达矩阵转化成列表
            with open('./dataSets/dataSets_filtered_unzipped/' + file_id + '_matrix.mtx', 'r') as f:
                data_lists = f.readlines()
                data_list = data_lists[2]
                data_list = data_list.strip('\n')
                data_list = data_list.split(' ')
                gene_num=data_list[0]
                cell_num=int(data_list[1])
                for i in range(1, cell_num + 1):
                    data_cell.append({'cell_id': str(i), 'cell_genes': [],'file_type':file_type})
                j = 0
                # for i in range(3,len(data_lists)):
                for data_list in data_lists:
                    if (j <= 2):
                        j += 1
                        continue
                    j += 1
                    data_list = data_list.strip('\n')
                    data_list = data_list.split(' ')
                    cell_gene = gene_id[data_list[0]]
                    cell_id = data_list[1]
                    data_cell[int(cell_id) - 1]['cell_genes'].append(cell_gene)
            cell_ids_filtered = []
            for i in range(len(data_cell)):
                if (len(data_cell[i]['cell_genes']) != 0):
                    cell_ids_filtered.append(i)
            for i in cell_ids_filtered:
                data_cells_all[file_type].append(data_cell[i])
            cells_num_all += len(data_cells_all[file_type])
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE34969'):
        data_files=['GSM859633.txt', 'GSM859626.txt', 'GSM859622.txt', 'GSM859625.txt', 'GSM859628.txt', 'GSM859630.txt', 'GSM859629.txt', 'GSM859635.txt', 'GSM859632.txt', 'GSM859624.txt', 'GSM859627.txt', 'GSM859631.txt', 'GSM859623.txt', 'GSM859634.txt']
        # GSM859622	6000 Rads 6hrs Replicate 1
        # GSM859623	6000 Rads 6hrs Replicate 2
        # GSM859624	6000 Rads 6hrs Replicate 3
        # GSM859625	6000 Rads 6hrs Replicate 4
        # GSM859626	6000 Rads 6hrs Replicate 5
        # GSM859627	6000 Rads 12hrs Replicate 1
        # GSM859628	6000 Rads 12hrs Replicate 2
        # GSM859629	6000 Rads 12hrs Replicate 3
        # GSM859630	6000 Rads 24hrs Replicate 1
        # GSM859631	6000 Rads 24hrs Replicate 2
        # GSM859632	6000 Rads 24hrs Replicate 3
        # GSM859633	6000 Rads 48hrs Replicate 1
        # GSM859634	6000 Rads 48hrs Replicate 2
        # GSM859635	6000 Rads 48hrs Replicate 3
        cells_num_all = 0
        data_cells_all={}
        for file_name in data_files:
            with open('./dataSets/dataSets_filtered_unzipped/'+file_name,'r')as f:
                data_lists=f.readlines()
                data_cell= {}
                cell_ids=[]
                for i in range(6,len(data_lists)):
                    data_list=data_lists[i]
                    data_list=data_list.strip('\n')
                    data_list=data_list.split('\t')
                    if(data_list[0]=='DATA' and ('SMED' in data_list[11])):
                        cell_id=data_list[2]
                        if(cell_id in data_cell.keys()):
                            data_cell[cell_id]['cell_genes'].append(data_list[11])
                        else:
                            data_cell[cell_id]={'cell_genes':[],'cell_id':cell_id,'file_type':file_name}
            data_cells=[]
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name]=data_cells
            cells_num_all+=len(data_cells)
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE35565'):
        filenames=['GSM870719_smed2rep3.gpr', 'GSM870717_smed2rep1.gpr', 'GSM870715_smed1rep3.gpr', 'GSM870713_smed1rep1.gpr', 'GSM870714_smed1rep2.gpr', 'GSM870716_smed1rep4.gpr', 'GSM870718_smed2rep2.gpr', 'GSM870720_smed2rep4.gpr']
        # GSM870713	肠道与对照 SmedArray1-Biorep1
        # GSM870714	肠道与对照 SmedArray1-Biorep2
        # GSM870715	肠道与对照 SmedArray1-Biorep3
        # GSM870716	肠道与对照 SmedArray1-Biorep4
        # GSM870717	肠道与对照 SmedArray2-Biorep1
        # GSM870718	肠道与对照 SmedArray2-Biorep2
        # GSM870719	肠道与对照 SmedArray2-Biorep3
        # GSM870720	肠道与对照 SmedArray2-Biorep4
        data_cells_all={}
        cells_num_all=0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/'+file_name,'r')as f:
                datalists=f.readlines()
                data_cell={}
                for i in range(33,len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    cell_id=datalist[2]
                    gene=datalist[3]
                    if (cell_id in data_cell.keys()):
                        data_cell[cell_id]['cell_genes'].append(gene)
                    else:
                        data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}
            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name]=data_cells
            cells_num_all+=len(data_cells)
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE36869'):
        filenames=['GSM904636_252175010035_12142010_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM904638_252175010032_12142010_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM904637_252175010034_12142010_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM904635_252175010030_201011021039_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM904638_252175010032_12142010_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM904637_252175010030_201011021039_S01_GE2-v5_10_Apr08_1_4.txt', 'GPL14150_021750_D_GEO_20081016.txt', 'GSM904635_252175010033_12142010_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM904638_252175010033_12142010_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM904636_252175010031_201011021047_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM904636_252175010034_12142010_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM904637_252175010031_201011021047_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM904635_252175010035_12142010_S01_GE2-v5_10_Apr08_1_4.txt']
        # GSM904635	5 分钟截肢的 runt-1 RNAi 动物（重复 1-3）
        # GSM904636	9h 截肢的 runt-1 RNAi 动物（重复 1-3）
        # GSM904637	24小时截肢runt-1 RNAi动物（重复1-3）
        # GSM904638	X1 新生细胞，从 9 小时截肢的 runt-1 RNAi 动物中分离出来（重复 1-3
        data_cells_all = {}
        cells_num_all = 0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/' + file_name, 'r') as f:
                datalists = f.readlines()
                data_cell = {}
                for i in range(1, len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    if(datalist[0]=='DATA'):
                        cell_id = datalist[2]
                        gene = datalist[11]
                        if('SMED'in gene):
                            if (cell_id in data_cell.keys()):
                                data_cell[cell_id]['cell_genes'].append(gene)
                            else:
                                data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}

            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name] = data_cells
            cells_num_all += len(data_cells)
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE36945'):
        filenames=['GSM906827_11.21.08_252175010007_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906827_11.21.08_252175010005_S01_GE2-v5_10_Apr08_1_1.txt', 'GPL14150_gene_models.txt', 'GSM906819_11.21.08_252175010002_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906822_11.21.08_252175010002_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906822_11.21.08_252175010004_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906824_11.21.08_252175010005_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906826_11.21.08_252175010008_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906819_11.21.08_252175010003_S01_GE2-v5_10_Apr08_1_1.txt', 'GSM906818_11.21.08_252175010002_S01_GE2-v5_10_Apr08_1_3.txt', 'GPL14150_021750_D_GEO_20081016.txt', 'GSM906826_11.21.08_252175010006_S01_GE2-v5_10_Apr08_1_1.txt', 'GSM906818_11.21.08_252175010003_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906825_11.21.08_252175010005_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906818_11.21.08_252175010001_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906820_11.21.08_252175010002_S01_GE2-v5_10_Apr08_1_1.txt', 'GSM906823_11.21.08_252175010007_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906820_11.21.08_252175010004_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906821_11.21.08_252175010003_S01_GE2-v5_10_Apr08_1_4.txt', 'GSM906826_11.21.08_252175010005_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906827_11.21.08_252175010008_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906821_11.21.08_252175010001_S01_GE2-v5_10_Apr08_1_1.txt', 'GSM906824_11.21.08_252175010006_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906820_11.21.08_252175010001_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906825_11.21.08_252175010006_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906825_11.21.08_252175010007_S01_GE2-v5_10_Apr08_1_1.txt', 'GSM906821_11.21.08_252175010004_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906823_11.21.08_252175010004_S01_GE2-v5_10_Apr08_1_1.txt', 'GSM906822_11.21.08_252175010003_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906824_11.21.08_252175010007_S01_GE2-v5_10_Apr08_1_2.txt', 'GSM906819_11.21.08_252175010001_S01_GE2-v5_10_Apr08_1_3.txt', 'GSM906823_11.21.08_252175010006_S01_GE2-v5_10_Apr08_1_4.txt']
        # GSM906818	30分钟截肢动物
        # GSM906819	1小时截肢动物
        # GSM906820	3小时截肢动物
        # GSM906821	6小时截肢动物
        # GSM906822	12小时截肢动物
        # GSM906823	30分钟受辐射并截肢的动物
        # GSM906824	1小时辐射和截肢动物
        # GSM906825	3小时辐射和截肢动物
        # GSM906826	6小时辐射和截肢动物
        # GSM906827	12小时辐射和截肢动物
        data_cells_all = {}
        cells_num_all = 0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/' + file_name, 'r') as f:
                datalists = f.readlines()
                data_cell = {}
                for i in range(1, len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    if(datalist[0]=='DATA'):
                        cell_id = datalist[2]
                        gene = datalist[11]
                        if('SMED'in gene):
                            if (cell_id in data_cell.keys()):
                                data_cell[cell_id]['cell_genes'].append(gene)
                            else:
                                data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}
            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name] = data_cells
            cells_num_all += len(data_cells)
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE56178'):
        filenames=['GSM1357197_US82600140_254527110004_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357214_US82600140_254527110008_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357211_US82600140_254527110007_S01_GE1_107_Sep09_red_only_1_4.txt', 'GSM1357213_US82600140_254527110006_S01_GE1_107_Sep09_red_only_1_4.txt', 'GSM1357202_US82600140_254527110007_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357195_US82600140_254527110003_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357204_US82600140_254527110005_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357194_US82600140_254527110002_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357206_US82600140_254527110004_S01_GE1_107_Sep09_red_only_1_4.txt', 'GSM1357196_US82600140_254527110007_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357209_US82600140_254527110001_S01_GE1_107_Sep09_red_only_1_3.txt', 'GSM1357212_US82600140_254527110005_S01_GE1_107_Sep09_red_only_1_4.txt', 'GSM1357210_US82600140_254527110002_S01_GE1_107_Sep09_red_only_1_3.txt', 'GSM1357192_US82600140_254527110002_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357205_US82600140_254527110006_S01_GE1_107_Sep09_red_only_1_3.txt', 'GSM1357200_US82600140_254527110001_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357208_US82600140_254527110006_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357193_US82600140_254527110004_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357201_US82600140_254527110003_S01_GE1_107_Sep09_red_only_1_2.txt', 'GSM1357198_US82600140_254527110006_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357191_US82600140_254527110001_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357203_US82600140_254527110003_S01_GE1_107_Sep09_red_only_1_3.txt', 'GSM1357199_US82600140_254527110008_S01_GE1_107_Sep09_red_only_1_1.txt', 'GSM1357207_US82600140_254527110005_S01_GE1_107_Sep09_red_only_1_3.txt']
        # GSM1357191	0h_A_X1(1)
        # GSM1357192	0h_A_X1(2)
        # GSM1357193	0h_A_X1(3)
        # GSM1357194	24h_A_X1(1)
        # GSM1357195	24h_A_X1(2)
        # GSM1357196	24小时_A_X1(3)
        # GSM1357197	48h_A_X1(1)
        # GSM1357198	48h_A_X1(2)
        # GSM1357199	48h_A_X1(3)
        # GSM1357200	72h_A_X1(1)
        # GSM1357201	72h_A_X1(2)
        # GSM1357202	72h_A_72(3)
        # GSM1357203	0h_P_X1(1)
        # GSM1357204	0h_P_X1(2)
        # GSM1357205	0h_P_X1(3)
        # GSM1357206	24h_P_X1(1)
        # GSM1357207	24h_P_X1(2)
        # GSM1357208	24h_P_X1(3)
        # GSM1357209	48h_P_X1(1)
        # GSM1357210	48h_P_X1(2)
        # GSM1357211	48h_P_X1(3)
        # GSM1357212	72h_P_X1(1)
        # GSM1357213	72h_P_X1(2)
        # GSM1357214	72h_P_X1(3
        data_cells_all = {}
        cells_num_all = 0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/' + file_name, 'r') as f:
                datalists = f.readlines()
                data_cell = {}
                for i in range(1, len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    if(datalist[0]=='DATA'):
                        cell_id = datalist[2]
                        gene = datalist[12]
                        if('SMED'in gene or 'CUST' in gene):
                            if (cell_id in data_cell.keys()):
                                data_cell[cell_id]['cell_genes'].append(gene)
                            else:
                                data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}
            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name] = data_cells
            cells_num_all += len(data_cells)
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE56181'):
        filenames=['GSM1357281_Control_24hr_2.txt', 'GSM1357284_Control_48hr_2.txt', 'GSM1357291_Irradiated_6hr_3.txt', 'GSM1357298_Irradiated_24hr_1.txt', 'GSM1357296_Irradiated_18hr_2.txt', 'GSM1357293_Irradiated_12hr_2.txt', 'GSM1357294_Irradiated_12hr_3.txt', 'GSM1357275_Control_12hr_2.txt', 'GSM1357292_Irradiated_12hr_1.txt', 'GSM1357297_Irradiated_18hr_3.txt', 'GSM1357300_Irradiated_24hr_3.txt', 'GSM1357303_Irradiated_48hr_3.txt', 'GSM1357272_Control_6hr_2.txt', 'GSM1357289_Irradiated_6hr_1.txt', 'GSM1357273_Control_6hr_3.txt', 'GSM1357282_Control_24hr_3.txt', 'GSM1357306_Irradiated_72hr_3.txt', 'GSM1357304_Irradiated_72hr_1.txt', 'GSM1357271_Control_6hr_1.txt', 'GSM1357280_Control_24hr_1.txt', 'GSM1357288_Control_72hr_3.txt', 'GSM1357305_Irradiated_72hr_2.txt', 'GSM1357295_Irradiated_18hr_1.txt', 'GSM1357279_Control_18hr_3.txt', 'GSM1357276_Control_12hr_3.txt', 'GSM1357285_Control_48hr_3.txt', 'GSM1357301_Irradiated_48hr_1.txt', 'GSM1357299_Irradiated_24hr_2.txt', 'GSM1357290_Irradiated_6hr_2.txt', 'GSM1357286_Control_72hr_1.txt', 'GSM1357277_Control_18hr_1.txt', 'GSM1357302_Irradiated_48hr_2.txt', 'GSM1357287_Control_72hr_2.txt', 'GSM1357274_Control_12hr_1.txt', 'GSM1357283_Control_48hr_1.txt', 'GSM1357278_Control_18hr_2.txt']
        # GSM1357271	截肢后 6 小时控制蠕虫咽部伤口塞重复 1
        # GSM1357272	截肢后 6 小时控制蠕虫咽伤口塞重复 2
        # GSM1357273	截肢后 6 小时控制蠕虫咽伤口塞重复 3
        # GSM1357274	截肢后 12 小时控制蠕虫咽伤口塞重复 1
        # GSM1357275	截肢后 12 小时控制蠕虫咽伤口塞重复 2
        # GSM1357276	截肢后 12 小时控制蠕虫咽伤口塞重复 3
        # GSM1357277	截肢后 18 小时控制蠕虫咽伤口塞重复 1
        # GSM1357278	截肢后 18 小时控制蠕虫咽伤口塞重复 2
        # GSM1357279	截肢后 18 小时控制蠕虫咽伤口塞重复 3
        # GSM1357280	截肢后 24 小时控制蠕虫咽伤口塞重复 1
        # GSM1357281	截肢后 24 小时控制蠕虫咽伤口塞重复 2
        # GSM1357282	截肢后 24 小时控制蠕虫咽伤口塞重复 3
        # GSM1357283	截肢后 48 小时控制蠕虫咽伤口塞重复 1
        # GSM1357284	截肢后 48 小时控制蠕虫咽伤口塞重复 2
        # GSM1357285	截肢后 48 小时控制蠕虫咽伤口塞重复 3
        # GSM1357286	截肢后 72 小时控制蠕虫咽伤口塞重复 1
        # GSM1357287	截肢后 72 小时控制蠕虫咽伤口塞重复 2
        # GSM1357288	截肢后 72 小时控制蠕虫咽伤口塞重复 3
        # GSM1357289	截肢后 6 小时辐照蠕虫咽部伤口塞重复 1
        # GSM1357290	截肢后 6 小时辐照蠕虫咽部伤口塞重复 2
        # GSM1357291	截肢后 6 小时辐照蠕虫咽伤口塞重复 3
        # GSM1357292	截肢后 12 小时辐照蠕虫咽部伤口塞重复 1
        # GSM1357293	截肢后 12 小时辐照蠕虫咽部伤口塞重复 2
        # GSM1357294	截肢后 12 小时辐照蠕虫咽部伤口塞重复 3
        # GSM1357295	截肢后 18 小时辐照蠕虫咽伤口塞重复 1
        # GSM1357296	截肢后 18 小时辐照蠕虫咽部伤口塞重复 2
        # GSM1357297	截肢后 18 小时辐照蠕虫咽伤口塞重复 3
        # GSM1357298	截肢后 24 小时辐照蠕虫咽伤口塞重复 1
        # GSM1357299	截肢后 24 小时辐照蠕虫咽伤口塞重复 2
        # GSM1357300	截肢后 24 小时辐照蠕虫咽伤口塞重复 3
        # GSM1357301	截肢后 48 小时辐照蠕虫咽伤口塞重复 1
        # GSM1357302	截肢后 48 小时辐照蠕虫咽部伤口塞重复 2
        # GSM1357303	截肢后 48 小时辐照蠕虫咽伤口塞重复 3
        # GSM1357304	截肢后 72 小时辐照蠕虫咽伤口塞重复 1
        # GSM1357305	截肢后 72 小时辐照蠕虫咽部伤口塞重复 2
        # GSM1357306	截肢后 72 小时辐照蠕虫咽伤口塞重复 3
        data_cells_all = {}
        cells_num_all = 0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/' + file_name, 'r') as f:
                datalists = f.readlines()
                data_cell = {}
                for i in range(1, len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    if(datalist[0]=='DATA'):
                        cell_id = datalist[2]
                        gene = datalist[12]
                        if('SMED'in gene):
                            if (cell_id in data_cell.keys()):
                                data_cell[cell_id]['cell_genes'].append(gene)
                            else:
                                data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}
            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name] = data_cells
            cells_num_all += len(data_cells)
            print(cells_num_all)
        print(gse_id + ' end with num:' + str(cells_num_all))
        return data_cells_all
    if (gse_id == 'GSE78161'):
        filenames=['GSM2068553_SG13134300_257257910001_S001_GE1_1105_Oct12_1_3.txt', 'GSM2068558_SG13134300_257257910001_S001_GE1_1105_Oct12_2_4.txt', 'GSM2068554_SG13134300_257257910001_S001_GE1_1105_Oct12_2_2.txt', 'GSM2068556_SG13134300_257257910001_S001_GE1_1105_Oct12_2_3.txt', 'GSM2068557_SG13134300_257257910001_S001_GE1_1105_Oct12_2_1.txt', 'GSM2068555_SG13134300_257257910001_S001_GE1_1105_Oct12_1_4.txt']
        # GSM2068553	RNA_X1_cells_Proximal_多聚腺苷酸化位点_Rep1
        # GSM2068554	RNA_X1_cells_Proximal_多聚腺苷酸化位点_Rep2
        # GSM2068555	RNA_X2_cells_Proximal_多聚腺苷酸化位点_Rep1
        # GSM2068556	RNA_X2_cells_Proximal_多聚腺苷酸化位点_Rep2
        # GSM2068557	RNA_Xins_cells_Proximal_多腺苷酸化位点_Rep1
        # GSM2068558	RNA_Xins_cells_Proximal_多腺苷酸化位点_Rep2
        data_cells_all = {}
        cells_num_all = 0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/' + file_name, 'r') as f:
                datalists = f.readlines()
                data_cell = {}
                for i in range(1, len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    if(datalist[0]=='DATA'):
                        cell_id = datalist[2]
                        gene = datalist[12]
                        if('Contig'in gene):
                            if (cell_id in data_cell.keys()):
                                data_cell[cell_id]['cell_genes'].append(gene)
                            else:
                                data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}
            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name] = data_cells
            cells_num_all += len(data_cells)
            print(cells_num_all)
        return data_cells_all
    if (gse_id == 'GSE78261'):
        filenames=['GSM2070580_SG13134300_257243610001_S001_GE1_1105_Oct12_1_4.txt', 'GSM2070581_SG13134300_257243610001_S001_GE1_1105_Oct12_2_3.txt', 'GSM2070582_SG13134300_257243610001_S001_GE1_1105_Oct12_2_1.txt', 'GSM2070578_SG13134300_257243610001_S001_GE1_1105_Oct12_1_3.txt', 'GSM2070583_SG13134300_257243610001_S001_GE1_1105_Oct12_2_4.txt', 'GSM2070579_SG13134300_257243610001_S001_GE1_1105_Oct12_2_2.txt']
        # GSM2070578	RNA_X1_cells_Distal_多聚腺苷酸化位点_Rep1
        # GSM2070579	RNA_X1_cells_Distall_聚腺苷酸化位点_Rep2
        # GSM2070580	RNA_X2_cells_Distal_多聚腺苷酸化位点_Rep1
        # GSM2070581	RNA_X2_cells_Distal_多聚腺苷酸化位点_Rep2
        # GSM2070582	RNA_Xins_cells_Distal_聚腺苷酸化位点_Rep1
        # GSM2070583	RNA_Xins_cells_Distal_多腺苷酸化位点_Rep2
        data_cells_all = {}
        cells_num_all = 0
        for file_name in filenames:
            with open('./dataSets/dataSets_filtered_unzipped/' + file_name, 'r') as f:
                datalists = f.readlines()
                data_cell = {}
                for i in range(1, len(datalists)):
                    datalist = datalists[i]
                    datalist = datalist.strip('\n')
                    datalist = datalist.split('\t')
                    if(datalist[0]=='DATA'):
                        cell_id = datalist[2]
                        gene = datalist[12]
                        if('Contig'in gene):
                            if (cell_id in data_cell.keys()):
                                data_cell[cell_id]['cell_genes'].append(gene)
                            else:
                                data_cell[cell_id] = {'cell_genes': [], 'cell_id': cell_id, 'file_type': file_name}
            data_cells = []
            for cell_id in data_cell.keys():
                data_cells.append(data_cell[cell_id])
            data_cells_all[file_name] = data_cells
            cells_num_all += len(data_cells)
            print(cells_num_all)
        return data_cells_all
    if(gse_id=='GSE256032'):
        filenames=['GSE256032_features.tsv', 'GSE256032_barcodes.tsv', 'GSE256032_matrix.mtx']
        data_cells_all=[]
        data_cells={}
        gene_id={}
        i=1
        # 获取基因id
        with open('./dataSets/dataSets_filtered_unzipped/GSE256032_features.tsv' ,'r') as f:
            data_lists=f.readlines()
            for data_list in data_lists:
                data_list=data_list.strip('\n')
                data_list=data_list.split('\t')
                gene_id[str(i)]=data_list[0]
                i+=1
        # 获取表达数据
        with open('./dataSets/dataSets_filtered_unzipped/GSE256032_matrix.mtx' ,'r') as f:
            data_lists=f.readlines()
            for i in range(2,len(data_lists)):
                data_list=data_lists[i]
                data_list=data_list.strip('\n')
                data_list=data_list.split(' ')
                gene = data_list[0]
                cell_id=data_list[1]
                if(cell_id in data_cells.keys()):
                    data_cells[cell_id].append(gene_id[gene])
                else:
                    data_cells[cell_id]=[gene_id[gene]]
        for id in data_cells.keys():
            data_cells_all.append({'cell_id':id,'cell_genes':data_cells[id]})
        print(len(data_cells_all))
        return data_cells_all
    if (gse_id == 'GSE246681'):
        filenames = ['GSE246681_barcodes.tsv', 'GSE246681_features.tsv', 'GSE246681_matrix.mtx']
        data_cells_all = []
        data_cells = {}
        gene_id = {}
        i = 1
        # 获取基因id
        with open('./dataSets/dataSets_filtered_unzipped/GSE246681_features.tsv', 'r') as f:
            data_lists = f.readlines()
            for data_list in data_lists:
                data_list = data_list.strip('\n')
                data_list = data_list.split('\t')
                gene_id[str(i)] = data_list[0]
                i += 1
        # 获取表达数据
        with open('./dataSets/dataSets_filtered_unzipped/GSE246681_matrix.mtx', 'r') as f:
            data_lists = f.readlines()
            for i in range(2, len(data_lists)):
                data_list = data_lists[i]
                data_list = data_list.strip('\n')
                data_list = data_list.split(' ')
                gene = data_list[0]
                cell_id = data_list[1]
                if (cell_id in data_cells.keys()):
                    data_cells[cell_id].append(gene_id[gene])
                else:
                    data_cells[cell_id] = [gene_id[gene]]
        for id in data_cells.keys():
            data_cells_all.append({'cell_id': id, 'cell_genes': data_cells[id]})
        print(len(data_cells_all))
        return data_cells_all

    pass
def ID2gene_pre():
    # 此程序用于将对应GSMID的GSM文件的原始数据中探针ID转化成gene名的预处理
    # 获取所有的GSE
    gse_ids=[]
    for root, dirs, files in os.walk('./dataSets/dataSets_filtered'):
        for dir in dirs:
            if('GSE' in dir):
                gse=dir.split('_')[0]
                gse_ids.append(gse)
    gse_files={}
    gse_files1 = {}
    for gse in gse_ids:
        gse_files[gse]=[]
        for root, dirs, files in os.walk('./dataSets/dataSets_filtered/'+gse+'_RAW'):
            gse_files1[gse]=files
    for gse in gse_ids:
        for root, dirs, files in os.walk('./dataSets/dataSets_filtered_unzipped'):
            for file in files:
                for file0 in  gse_files1[gse]:
                    if(file in file0):
                        gse_files[gse].append(file)
                    if(gse in file): # 除.gz格式外也有其他命名格式
                        gse_files[gse].append(file)
        gse_files[gse]=list(set(gse_files[gse])) # 去除重复
    # print(gse_ids)
    gse_ids=list(set(gse_ids))
    # 所有GSE对应的数据文件已获取完毕
    # 对每个GSE的源数据分别处理
    gseAgpl={'GSE256032':'GPL29967','GSE246681':'GPL29967','GSE100818':'GPL23660', 'GSE103633':'GPL21689', 'GSE146685':'GPL21689', 'GSE150259':'GPL28515', 'GSE157860': 'GPL20150', 'GSE158706':'GPL20150', 'GSE179290':'GPL15553',
    'GSE179295':'GPL15553','GSE192524':'GPL21689', 'GSE212136':'GPL21689', 'GSE22797':'GPL10652', 'GSE231548':'GPL33372', 'GSE32450':'GPL10652', 'GSE34326':'GPL14991',
    'GSE34969':'GPL14150', 'GSE35565':'GPL15192|GPL15193','GSE36869':'GPL14150','GSE36945':'GPL14150', 'GSE56178':'GPL18464', 'GSE56181':'GPL18465', 'GSE62551':'GPL10652',
    'GSE75594':'GPL15553', 'GSE78161':'GPL21495','GSE78261':'GPL21512', 'GSE11503':'GPL6871'}
    gse_files['GSE157860']=['GSE157860_RNA-low_scRNAseq_data.xlsx'] # GSE157860就这一个文件可以用
    # print(len(gse_files))
    gse_valid=['GSE179290','GSE231548','GSE34969','GSE35565','GSE36869','GSE36945','GSE56178','GSE56181']
    # print(gse_files['GSE246681'])
    for gse in gse_valid:
        print(gse_files[gse])
        data_cell_all=datatreat(gse)
        with open('./data_treated_sets/'+gse+'.pkl','wb')as f:
            pickle.dump(data_cell_all,f)
            print(gse+'success')
    pass

def getIdDict(pathway,id=0):
    # 读取不同版本基因ID的匹配文件
    id_dict={}
    if(id==2):
        workbook = openpyxl.load_workbook(pathway)
        worksheet = workbook.active
        for row in worksheet.iter_rows():
            row_data=[cell.value for cell in row]
            if(row_data[0] in id_dict):
                continue
            id_dict[row_data[0]]=row_data[1]
        return id_dict
    with open(pathway, 'r') as f:
        data_lines=f.readlines()
        for data_line in data_lines:
            data_line=data_line.strip('\n')
            data_line=data_line.split('\t')
            gene_id0=data_line[0]
            if(id==0):
                index = gene_id0.rfind('_')
                gene_id0=gene_id0[:index]
            gene_id1=data_line[1]
            index = gene_id1.rfind('.')
            gene_id1 = gene_id1[:index]
            if gene_id0 in id_dict.keys():
                # 仅取s值最高的
                continue
            else:
                id_dict[gene_id0]=gene_id1
    # print(id_dict)
    return id_dict
    pass
def getSmest(pathway):
    id_dict={}
    if(id==2):
        with open(pathway, 'r') as f:
            data_lines=f.readlines()
            print(data_lines)
    with open(pathway, 'r') as f:
        data_lines = f.readlines()
        for data_line in data_lines:
            data_line = data_line.strip('\n')
            data_line = data_line.split('\t')
            gene_id0 = data_line[0]
            index = gene_id0.rfind('_')
            gene_id0 = gene_id0[:index]
            gene_id1 = data_line[1]
            index = gene_id0.rfind('.')
            gene_id1 = gene_id1[:index]
            if gene_id0 in id_dict.keys():
                # 仅取s值最高的
                continue
            else:
                id_dict[gene_id1] = gene_id0
    # print(id_dict)
    return id_dict
# getIdDict('./dataSets/result.txt')
def view_bar(message, num, total):
    # 进度条
    rate = num / total
    rate_num = int(rate * 30)
    rate_nums = math.ceil(rate * 100)
    r = '\r%s:[%s%s]%d%%\t%d/%d' % (message, "|" * rate_num, " " * (30 - rate_num), rate_nums, num, total,)
    sys.stdout.write(r)
    sys.stdout.flush()
def id2gene():
    # 将所有gse得到的细胞数据的基因名称统一
    # 有效的gse列表
    gseAgpl = {'GSE100818': 'GPL23660', 'GSE103633': 'GPL21689', 'GSE146685': 'GPL21689', 'GSE150259': 'GPL28515',
               'GSE157860': 'GPL20150', 'GSE158706': 'GPL20150', 'GSE179290': 'GPL15553',
               'GSE179295': 'GPL15553', 'GSE192524': 'GPL21689', 'GSE212136': 'GPL21689', 'GSE22797': 'GPL10652',
               'GSE231548': 'GPL33372', 'GSE32450': 'GPL10652', 'GSE34326': 'GPL14991',
               'GSE34969': 'GPL14150', 'GSE35565': 'GPL15192|GPL15193', 'GSE36869': 'GPL14150', 'GSE36945': 'GPL14150',
               'GSE56178': 'GPL18464', 'GSE56181': 'GPL18465', 'GSE62551': 'GPL10652',
               'GSE75594': 'GPL15553', 'GSE78161': 'GPL21495', 'GSE78261': 'GPL21512', 'GSE11503': 'GPL6871'}
    gse_valid = ['GSE179290', 'GSE231548', 'GSE256032','GSE246681','GSE34969', 'GSE56178', 'GSE56181','GSE35565', 'GSE36869', 'GSE36945']

    # 依次获取各个gse的基因表达数据，并将其转化成SMESG命名
    # 其中GSE179290 是 dd_Smed_v6，GSE231548'是SMESG，GSE256032,GSE246681是h1SMcG， 为主要数据集
    id_dict=getIdDict('./dataSets/result.txt')
    # 'GSE179290' 奇怪的bug, data_cell_all不能直接转dataset，但可以分步转。。
    for id in gse_valid:
        if(id=='GSE179290'):
            # continue
            id_dict = getIdDict('./dataSets/result.txt')
            print(len(id_dict))
            data_cell_all=datatreat(id)
            with open('./data_treated_sets/'+id+'.pkl', 'wb') as f:
                pickle.dump(data_cell_all, f)
                print('saved success')
            with open('./data_treated_sets/'+id+'.pkl', 'rb') as f:
                dataset_list=pickle.load(f)
            def toSmesg(example):
                example['cell_genes']=[id_dict[item] for item in example['cell_genes'] if item in id_dict.keys()]
                return example
            dataset_list2={}
            for dataset_name in dataset_list:
                dataset_item=dataset_list[dataset_name]
                # id转换
                dataset_item=dataset_item.map(toSmesg,num_proc=16)
                # # 去除表达较少的细胞
                # dataset_item=dataset_item.filter(if_not_short,num_proc=16)
                print(dataset_item)
                dataset_list2[dataset_name]=dataset_item
            with open('./data_treated_sets/'+id+'_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list2,f)
            print(id+' mapped success')
        if(id=='GSE231548'):
            # continue
            dataset_list=datatreat(id)
            with open('./data_treated_sets/' + id + '.pkl', 'wb') as f:
                pickle.dump(dataset_list, f)
            def toSmesg2(example):
                example['cell_genes']=[item for item in example['cell_genes'] if 'SMESG' in item]
                return example
            with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
                data_cell_list = pickle.load(f)
                dataset_list={}
                for name in data_cell_list.keys():
                    dataset_item=Dataset.from_list(data_cell_list[name])
                    dataset_item=dataset_item.map(toSmesg2,num_proc=16)
                    dataset_list[name]=dataset_item
            print(dataset_list)
            with open('./data_treated_sets/'+id+'_mapped.pkl', 'wb') as f:
                 pickle.dump(dataset_list,f)
            print(id + ' mapped success')
        if(id=='GSE256032'):
            # h1SMcG 的基因序列fasta文件没有找到，暂时跳过
            # data_cell_list=datatreat(id)
            # print(data_cell_list)
            # continue
            with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
                data_cell_list=pickle.load(f)
                print(data_cell_list)
        if(id=='GSE246681'):
            # h1SMc
            # continue
            data_cell_list=datatreat(id)
            with open('./data_treated_sets/' + id + '.pkl', 'wb') as f:
                pickle.dump(data_cell_list,f)
            # 此数据为列表
            continue
        if(id=='GSE34969'):
            # with open('./data_treated_sets/'+id+'_mapped.pkl', 'rb') as f:
            #     dataset_list=pickle.load(f)
            #     for item in dataset_list['GSM859626.txt']:
            #         print(item)
            #     print(dataset_list['GSM859626.txt'])
            # continue
            id_dict=getIdDict('./dataSets/result2.txt',1)
            def toSmesg3(example):
                example['cell_genes'] = [id_dict[item] for item in example['cell_genes'] if item in id_dict.keys()]
                return example
            with open('./data_treated_sets/' + id + '.pkl','rb')as f:
                data_cell_list=pickle.load(f)
                dataset_list = {}
                for name in data_cell_list.keys():
                    dataset_item = Dataset.from_list(data_cell_list[name])
                    dataset_item = dataset_item.map(toSmesg3, num_proc=16)
                    dataset_list[name] = dataset_item
            with open('./data_treated_sets/'+id+'_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list,f)
            print(id + ' mapped success')
        if(id=='GSE56178'):
            # continue
            id_dict=getIdDict('./dataSets/result3.txt')
            id_dict2 = getIdDict('./dataSets/result2.txt',1)
            def toSmesg4(example):
                items=[]
                for gene_id in example['cell_genes']:
                    index = gene_id.rfind('_')
                    gene_id = gene_id[:index]
                    items.append(gene_id)
                cell_genes=[]
                for gene_id in example['cell_genes']:
                    if(gene_id in id_dict2.keys()):
                        cell_genes.append(id_dict2[gene_id])
                    elif(gene_id in id_dict.keys()):
                        cell_genes.append(id_dict[gene_id])
                example['cell_genes'] = cell_genes
                return example
            with open('./data_treated_sets/' + id + '.pkl','rb')as f:
                data_cell_list=pickle.load(f)
                dataset_list = {}
                # print(data_cell_list)
                for name in data_cell_list.keys():
                    dataset_item = Dataset.from_list(data_cell_list[name])
                    dataset_item = dataset_item.map(toSmesg4, num_proc=16)
                    dataset_list[name] = dataset_item
            print(dataset_list)
            with open('./data_treated_sets/' + id + '_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list, f)
            pass
        if(id=='GSE56181'):
            # continue
            # 用的是SMED
            id_dict = getIdDict('./dataSets/result5.txt',1)

            #  'SMED_10099_V2_1' SMED10047386
            def toSmesg5(example):
                example['cell_genes'] = [id_dict[item] for item in example['cell_genes'] if item in id_dict.keys()]
                return example
            with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
                data_cell_list = pickle.load(f)
                dataset_list = {}
                for name in data_cell_list.keys():
                    dataset_item = Dataset.from_list(data_cell_list[name])
                    dataset_item = dataset_item.map(toSmesg5, num_proc=16)
                    dataset_list[name] = dataset_item

            print(dataset_list)
            with open('./data_treated_sets/' + id + '_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list, f)
        if(id=='GSE35565'):
            # continue
            id_dict = getIdDict('./dataSets/result6.txt',1)
            # print(id_dict)
            def toSmesg6(example):
                cell_genes=[]
                for item in example['cell_genes']:
                    item=item.split('"')
                    item=item[1]
                    cell_genes.append(item)
                example['cell_genes'] = [id_dict[item] for item in cell_genes if item in id_dict.keys()]
                return example

            with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
                data_cell_list = pickle.load(f)
                dataset_list = {}
                # print(data_cell_list)

                for name in data_cell_list.keys():
                    data_cell_list[name]
                    dataset_item = Dataset.from_list(data_cell_list[name])
                    dataset_item = dataset_item.map(toSmesg6, num_proc=16)
                    dataset_list[name] = dataset_item
            print(dataset_list)
            with open('./data_treated_sets/' + id + '_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list, f)
        if (id == 'GSE36869'):
            # continue

            id_dict = getIdDict('./dataSets/result2.txt',1)
            def toSmesg6(example):
                example['cell_genes'] = [id_dict[item] for item in example['cell_genes'] if item in id_dict.keys()]
                return example
            with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
                data_cell_list = pickle.load(f)
                dataset_list = {}
                # print(data_cell_list)
                for name in data_cell_list.keys():
                    dataset_item = Dataset.from_list(data_cell_list[name])
                    dataset_item = dataset_item.map(toSmesg6, num_proc=16)
                    dataset_list[name] = dataset_item
            print(dataset_list)
            with open('./data_treated_sets/' + id + '_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list, f)
        if (id == 'GSE36945'):
            # with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
            #     data_cell_list = pickle.load(f)
            #     print(data_cell_list)
            id_dict = getIdDict('./dataSets/result2.txt',1)
            def toSmesg6(example):
                example['cell_genes'] = [id_dict[item] for item in example['cell_genes'] if item in id_dict.keys()]
                return example

            with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
                data_cell_list = pickle.load(f)
                dataset_list = {}
                # print(data_cell_list)
                for name in data_cell_list.keys():
                    dataset_item = Dataset.from_list(data_cell_list[name])
                    dataset_item = dataset_item.map(toSmesg6, num_proc=16)
                    dataset_list[name] = dataset_item
            print(dataset_list)
            with open('./data_treated_sets/' + id + '_mapped.pkl', 'wb') as f:
                pickle.dump(dataset_list, f)
#
# id2gene()
def makeFASTA():
    # 将GPL14150_gene_models.txt内容转化成fasta文件
    all_data=[]
    with open('./dataSets/dataSets_filtered_unzipped/GPL14150_gene_models.txt','r')as f:
        datalists=f.readlines()
        for datalist in datalists:
            datalist=datalist.strip('\n')
            datalist=datalist.split('\t')
            gene_name=datalist[0]
            gene_seq=datalist[1]
            all_data.append([gene_name,gene_seq])
    with open('./dataSets/SMED_V2.fasta','w') as f:
        for item in all_data:
            f.write('>'+item[0]+'\n'+item[1]+'\n')

    # with open('dataSets/dd_Smed_v6.pcf.contigs.fasta/dd_Smed_v6.pcf.contigs.fasta','r') as f:
    #     data=f.readlines()
    #     print(data)
# makeFASTA()
def makeFASTA2():
    # 将GPL18464-16222.txt内容转化成fasta文件
    all_data=[]
    with open('./dataSets/dataSets_filtered_unzipped/GPL18464-16222.txt','r')as f:
        datalists=f.readlines()
        for i in range(22,len(datalists)):
            datalist = datalists[i]
            datalist=datalist.strip('\n')
            datalist=datalist.split('\t')
            gene_name=datalist[1]
            gene_seq=datalist[5]
            if(gene_seq):
                all_data.append([gene_name,gene_seq])
    with open('./dataSets/SMED_CUSTS.fasta','w') as f:
        for item in all_data:
            f.write('>'+item[0]+'\n'+item[1]+'\n')

    # with open('dataSets/dd_Smed_v6.pcf.contigs.fasta/dd_Smed_v6.pcf.contigs.fasta','r') as f:
    #     data=f.readlines()
    #     print(data)
# makeFASTA2()
def makeFASTA3():
    print('111')
    all_data=[]
    # with open('./dataSets/GPL_map_files/GPL18464_map.csv', 'r') as f:
    #     datalists = f.readlines()
    #     for i in range(len(datalists)):
    #         datalist = datalists[i]
    #         datalist = datalist.strip('\n')
    #         datalist = datalist.split(',')
    #         try:
    #             gene_name = datalist[1]
    #             gene_seq = datalist[5]
    #         except:
    #             continue
    #         if (gene_seq):
    #             all_data.append([gene_name, gene_seq])
    # print('111')
    with open('./dataSets/GPL_map_files/GPL18465_map.csv', 'r') as f:
        datalists = f.readlines()
        for i in range(len(datalists)):
            datalist = datalists[i]
            datalist = datalist.strip('\n')
            datalist = datalist.split(',')
            try:
                gene_name = datalist[1]
                gene_seq = datalist[7]
            except:
                continue
            if (gene_seq):
                all_data.append([gene_name, gene_seq])
    with open('./dataSets/SMED_2.fasta', 'w') as f:
        for item in all_data:
            f.write('>' + item[0] + '\n' + item[1] + '\n')
    print('111')
    pass
# makeFASTA3()
def makeFASTA4():
    print('111')
    all_data=[]
    with open('./dataSets/GPL_map_files/GPL15192_map.csv', 'r') as f:
        datalists = f.readlines()
        for i in range(len(datalists)):
            datalist = datalists[i]
            datalist = datalist.strip('\n')
            datalist = datalist.split(',')
            try:
                gene_name = datalist[0]
                gene_seq = datalist[1]
            except:
                continue
            if (gene_seq):
                all_data.append([gene_name, gene_seq])
    print('111')
    with open('./dataSets/GPL_map_files/GPL15193_map.csv', 'r') as f:
        datalists = f.readlines()
        for i in range(len(datalists)):
            datalist = datalists[i]
            datalist = datalist.strip('\n')
            datalist = datalist.split(',')
            try:
                gene_name = datalist[0]
                gene_seq = datalist[1]
            except:
                continue
            if (gene_seq):
                all_data.append([gene_name, gene_seq])
    with open('./dataSets/Config1.fasta', 'w') as f:
        for item in all_data:
            f.write('>' + item[0] + '\n' + item[1] + '\n')
    print('111')
    pass
# makeFASTA4()
def getOther_cells_all():
    # 将其他数据集的涡虫细胞数据去标签合并，并将其进行预处理，为之后的无监督学习做准备
    data_cells_all=[]
    # 其余数据集的细胞主要集中于'GSE179290', 'GSE231548'这两个数据集中 此外还有'GSE34969' ，其余数据集的匹配文件没找到,且数据质量不高
    dataSets=['GSE179290', 'GSE231548','GSE34969','GSE56178','GSE56181','GSE35565', 'GSE36869', 'GSE36945']
    i = 0
    gene_len = 10
    for id in dataSets:
        k=0
        with open('./data_treated_sets/' + id + '.pkl', 'rb') as f:
            data_cells = pickle.load(f)
            for type in data_cells.keys():
                cell_lists = data_cells[type]
                for cell in cell_lists:
                    cell_item = {}
                    cell_item['cell_genes'] = cell['cell_genes']
                    cell_item['length'] = len(cell['cell_genes'])
                    if (cell_item['length'] < gene_len):
                        # 基因数据太少的细胞忽略
                        k += 1
                        continue
                    i += 1
                    view_bar('Processing cell unmapped', i, 500000)
        print('\n' + ' unmapped ' + id + '  ' + str(k))
    i = 0
    nums={}
    sum=0
    for id in dataSets:
        k=0
        k2=0
        with open('./data_treated_sets/'+id+'_mapped.pkl', 'rb') as f:
            data_cells=pickle.load(f)
            for type in data_cells.keys():
                cell_lists=data_cells[type]
                for cell in cell_lists:
                    cell_item={}
                    cell_item['cell_genes']=cell['cell_genes']
                    random.shuffle(cell_item['cell_genes'])
                    cell_item['length']=len(cell['cell_genes'])
                    if(cell_item['length']<gene_len):
                        # 基因数据太少的细胞忽略
                        k+=1
                        continue
                    cell_item['cell_id'] = str(i)
                    data_cells_all.append(cell_item)
                    i+=1
                    k2+=1
                    view_bar('Processing cell ',i,500000)
        print('\n'+' mapped '+id+'  '+str(k))
        nums[id]=k2
        sum+=k2
    for id in nums:
        print(id + '_num_'+str(float(nums[id])/sum))
    with open('./data_treated_sets/other_valid_p_cell/p_cell_unannotated.pkl', 'wb') as f:
        pickle.dump(data_cells_all,f)
    print('sucess')
# id2gene()
# gzip_all('./dataSets/dataSets_filtered')
# ID2gene_pre()
# datatreat('GSE11503','GSM289258.txt') # 可用10
# datatreat('GSE100818','GSM2693971_Fi02002.txt'
# datatreat('GSE103633','GSM2777019_x1_DGE_CLEAN.MULTI.txt')
# datatreat('GSE146685','GSM4404049_UI_D07_expression_matrix.csv')
# datatreat('GSE150259','GSM4544459_Smed_L1_MQ0_MG100_int.txt')
# datatreat('GSE157860','GSE157860_RNA-low_scRNAseq_data.xlsx') 可用80
# datatreat('GSE158706','GSM4808633_sirNeoblast-cells-10x.csv')
# datatreat('GSE179290') # 可用2484036
# datatreat('GSE212136') # 基因数据检测，与本实验无关
# datatreat('GSE231548') # 可用1120377
# datatreat('GSE34969') # 可用7448
# datatreat('GSE35565')  # 可用1792
# datatreat('GSE36869')# 可用6384
# datatreat('GSE36945') # 可用 159600
# datatreat('GSE56178') # 可用 25536
# datatreat('GSE56181') # 可用19152
# datatreat('GSE78161') # 1152 考虑实验性质，不可用
# datatreat('GSE78261')# 1152 考虑实验性质，不可用
# datatreat('GSE256032') # 可用95420
# datatreat('GSE246681') # 可用38576
# getOther_cells_all()
